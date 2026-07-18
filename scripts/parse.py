"""Extrahiert Items aus data/raw/ (HTML-Tabs + xlsx) und data/manual/ (xlsx) nach site/items.json.

Heuristik: Zelle mit Kauf-Link = Item-Anker; Name/Preis/Bild aus Nachbarzellen
derselben Zeile; Kategorie aus Tab-Name, Abschnitts-Headern und Namens-Keywords.
"""
import html as htmllib
import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

import openpyxl
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
MANUAL = ROOT / "data" / "manual"
OUT = ROOT / "site" / "items.json"

# Kauf-Link-Domains (Agenten + Marktplaetze)
SHOP_HOSTS = (
    "weidian.com", "taobao.com", "tmall.com", "1688.com", "yupoo.com",
    "kakobuy.com", "sugargoo.com", "cnfans.com", "mulebuy.com", "acbuy.com",
    "superbuy.com", "cssbuy.com", "hipobuy.com", "joyagoo.com", "pandabuy.com",
    "hagobuy.com", "allchinabuy.com", "oopbuy.com", "orientdig.com",
    "hoobuy.com", "loongbuy.com", "ponybuy.com", "basetao.com", "wegobuy.com",
    "itaobuy.com", "sifubuy.com", "eastmallbuy.com", "usfans.com", "lovegobuy.com",
    "ootdbuy.com", "panglobalbuy.com", "bbdbuy.com", "gjbuy.com", "vigorbuy.com",
)

PRICE_RE = re.compile(
    r"(?:[$€£¥￥]\s?\d[\d.,]*|\d[\d.,]*\s?(?:USD|EUR|CNY|RMB|GBP|\$|€|¥|￥)|(?:USD|EUR|CNY|RMB|GBP)\s?\d[\d.,]*)",
    re.I,
)
HYPERLINK_RE = re.compile(r'HYPERLINK\(\s*""?"?([^"]+?)""?"?\s*[,)]', re.I)
IMAGE_RE = re.compile(r'IMAGE\(\s*""?"?(https?://[^"]+?)""?"?\s*[,)]', re.I)
URL_RE = re.compile(r"https?://[^\s\"'<>)]+")
# Video-Thumbnails und UI-Icons sind keine Produktbilder
BAD_IMG_RE = re.compile(r"ytimg\.com|youtube\.com|youtu\.be|gstatic\.com/docs|/branding/", re.I)


def clean_img(url):
    if not url or BAD_IMG_RE.search(url):
        return None
    return url

AGENTS = (
    "kakobuy|sugargoo|cnfans|mulebuy|acbuy|superbuy|cssbuy|hipobuy|joyagoo|pandabuy|"
    "hagobuy|allchinabuy|oopbuy|orientdig|hoobuy|loongbuy|ponybuy|basetao|wegobuy|"
    "itaobuy|sifubuy|eastmallbuy|usfans|lovegobuy|ootdbuy|panglobalbuy|bbdbuy|gjbuy|"
    "vigorbuy|weidian|taobao|tmall|1688|yupoo"
)
BAD_NAME = re.compile(
    r"^(link|links|click|click here|here|raw link|buy|photo|photos|picture|pic|image|img|qc|view|"
    r"name|names|item name|price|prices|\W*|\w+ links?|(" + AGENTS + r")( links?)?)$",
    re.I,
)
NOISE_NAME = re.compile(
    r"discord|instagram|tiktok|youtube|reddit|sign up|coupon|tutorial|how to|join|follow|subscribe|free \$|% off|click here|spreadsheet|ctrl\s*\+\s*f|telegram|whatsapp|converter",
    re.I,
)

# Eindeutige Produkttyp-Woerter schlagen Marken-Proxys und Tab-Namen
STRONG_TYPES = [
    ("Shirts & Tees", r"\btees?\b|t-?shirts?\b|\bshirts?\b|\bpolos?\b|longsleeve|long sleeve|\bvests?\b|tank top"),
    ("Hoodies & Sweater", r"hoodies?\b|sweaters?\b|sweatshirts?\b|crewnecks?\b|zip.?ups?\b|cardigans?\b|pullovers?\b|fleece\b"),
    ("Jacken", r"jackets?\b|puffers?\b|\bcoats?\b|parkas?\b|windbreakers?\b|varsity|bombers?\b"),
    ("Hosen & Shorts", r"\bpants\b|\bjeans\b|\bshorts\b|sweatpants?\b|trousers?\b|cargos?\b|tracksuits?\b|joggers?\b|boxers?\b"),
    ("Trikots", r"jerseys?\b|trikots?\b"),
    ("Taschen", r"\bbags?\b|backpacks?\b|totes?\b|crossbody|messengers?\b|duffle|wallets?\b|cardholders?\b"),
    ("Uhren", r"\bwatch\b|\bwatches\b"),
    ("Schuhe", r"\bshoes?\b|sneakers?\b|\bslides?\b|slippers?\b|sandals?\b|\bboots?\b|loafers?\b|trainers?\b|\bdunks?\b|\bheels?\b|\bcleats?\b"),
    ("Parfum", r"parfums?\b|perfumes?\b|colognes?\b|fragrances?\b"),
    ("Schmuck & Accessoires", r"necklaces?\b|pendants?\b|bracelets?\b|earrings?\b|\bbelts?\b|\bcaps?\b|beanies?\b|sunglass|\bsocks?\b|\brings?\b|scarf|scarves"),
]
STRONG_TYPES_C = [(c, re.compile(p, re.I)) for c, p in STRONG_TYPES]

CATEGORIES = [
    ("Schuhe", r"shoe|sneaker|slide|slipper|sandal|boot|loafer|trainer|dunk|jordan|\baj\d|af1|airmax|air max|yeezy|new balance|bapesta|\bsb\b|foam|croc|heel|mule|birkenstock|samba|gazelle|campus|3xl|b30|b22|b27|tabi"),
    ("Trikots", r"jersey|jerseys|trikot|#\d+ |city edition|nba|nfl|soccer|football kit"),
    ("Shirts & Tees", r"\btee\b|tees|t-?shirt|shirt|polo|longsleeve|long sleeve|vest|tank"),
    ("Hoodies & Sweater", r"hoodie|sweater|sweatshirt|crewneck|zip|cardigan|knit|fleece|pullover"),
    ("Jacken", r"jacket|jackets|puffer|coat|parka|windbreaker|varsity|denim jacket|bomber|monclair|moncler"),
    ("Hosen & Shorts", r"pant|pants|jean|jeans|short|shorts|sweatpant|trouser|cargo|tracksuit|track suit|leggings|boxers|underwear|undies"),
    ("Taschen", r"\bbag\b|bags|backpack|tote|crossbody|messenger|duffle|wallet|cardholder|card holder|pouch|carpenter"),
    ("Uhren", r"watch|watches|rolex|relox|royal oak|richard|patek|omega|cartier santos|datejust|submariner|nautilus"),
    ("Schmuck & Accessoires", r"jewel|rings?\b|necklace|pendant|chain|bracelet|earring|belt|caps?\b|hats?\b|beanie|sunglass|glasses|scarf|glove|sock|accessor|keychain|airpods case|phone case"),
    ("Parfum", r"parfum|perfume|cologne|fragrance|tom ford|dior sauvage|creed"),
    ("Elektronik", r"electronic|airpods|iphone|ipad|samsung|dyson|bose|\bjbl\b|sony|beats|marshall|jabra|headphone|earbud|speaker|playstation|controller|lego|philips|shure|console|smartwatch"),
]


BRANDS = [
    ("Polo Ralph Lauren", r"polo ralph lauren|ralph lauren|\brl\b|big pony"),
    ("Louis Vuitton", r"louis vuitton|\blv\b|louiis"),
    ("Stone Island", r"stone island"),
    ("CP Company", r"c\.?p\.? company"),
    ("Fear of God", r"fear of god|essentials|\bfog\b"),
    ("The North Face", r"north face|\btnf\b"),
    ("Chrome Hearts", r"chrome hearts?"),
    ("Vivienne Westwood", r"vivienne|westwood"),
    ("New Balance", r"new balance|\bnb\d{3,4}\b"),
    ("LEGO", r"\blego\b"),
    ("Maison Margiela", r"margiela|\bmm6\b|tabi"),
    ("Canada Goose", r"canada goose"),
    ("Dr. Martens", r"dr\.? ?martens"),
    ("Palm Angels", r"palm angels"),
    ("Denim Tears", r"denim tears"),
    ("Syna World", r"syna ?world"),
    ("Off-White", r"off.?white"),
    ("Rick Owens", r"rick owens"),
    ("Comme des Garcons", r"comme des|\bcdg\b"),
    ("Audemars Piguet", r"audemars|royal oak|\bap\b"),
    ("Patek Philippe", r"patek|nautilus"),
    ("Richard Mille", r"richard mille|richards?\b"),
    ("Rolex", r"rolex|relox|daytona|datejust|submariner"),
    ("Tom Ford", r"tom ford"),
    ("Bottega Veneta", r"bottega"),
    ("Saint Laurent", r"saint laurent|\bysl\b"),
    ("Michael Kors", r"michael kors"),
    ("Tommy Hilfiger", r"tommy"),
    ("Calvin Klein", r"calvin klein|\bck\b"),
    ("Jordan", r"jordan|\baj\d|\bj\d{1,2}s?\b"),
    ("Nike", r"\bnike\b|air force|\baf1\b|air ?max|dunk|nocta|techfleece|tech fleece|shox"),
    ("Adidas", r"adidas|samba|gazelle|campus|superstar"),
    ("Yeezy", r"yeezy"),
    ("Gucci", r"gucci"),
    ("Dior", r"\bdior\b|b30|b22|b23|b27"),
    ("Chanel", r"chanel"),
    ("Prada", r"prada"),
    ("Balenciaga", r"balenciaga|balenci|\bblcg\b|3xl"),
    ("Corteiz", r"corteiz|\bcrtz\b"),
    ("Trapstar", r"trapstar"),
    ("Supreme", r"supreme"),
    ("BAPE", r"\bbape\b|bapesta|bathing ape"),
    ("Moncler", r"moncler|monclair"),
    ("Carhartt", r"carhartt"),
    ("Stussy", r"stussy|stüssy"),
    ("Amiri", r"amiri"),
    ("Burberry", r"burberry"),
    ("Hellstar", r"hellstar"),
    ("Sp5der", r"sp5der|spider hoodie"),
    ("Casablanca", r"casablanca"),
    ("Arc'teryx", r"arc.?teryx"),
    ("Patagonia", r"patagonia"),
    ("Asics", r"asics"),
    ("UGG", r"\bugg\b"),
    ("Timberland", r"timberland"),
    ("Birkenstock", r"birkenstock"),
    ("Converse", r"converse|chuck taylor"),
    ("Vans", r"\bvans\b"),
    ("Omega", r"\bomega\b|seamaster|speedmaster"),
    ("Cartier", r"cartier|santos"),
    ("Versace", r"versace"),
    ("Fendi", r"fendi"),
    ("Hermes", r"hermes|birkin"),
    ("Goyard", r"goyard"),
    ("Celine", r"celine"),
    ("Loewe", r"loewe"),
    ("Miu Miu", r"miu miu"),
    ("Armani", r"armani"),
    ("Hugo Boss", r"hugo boss|\bboss\b"),
    ("Lacoste", r"lacoste"),
    ("Lululemon", r"lulu"),
    ("Under Armour", r"under armou?r"),
    ("Puma", r"\bpuma\b"),
    ("Salomon", r"salomon"),
    ("On Running", r"on cloud|on running"),
    ("Hoka", r"\bhoka\b"),
    ("Apple", r"\bapple\b|iphone|ipad|airpods|apple watch"),
    ("Dyson", r"dyson"),
    ("Bose", r"\bbose\b"),
    ("JBL", r"\bjbl\b"),
    ("Sony", r"\bsony\b|playstation"),
    ("Samsung", r"samsung"),
    ("Beats", r"\bbeats\b"),
    ("Marshall", r"marshall"),
]
BRANDS_C = [(name, re.compile(pat, re.I)) for name, pat in BRANDS]


def brand_of(name):
    for b, rx in BRANDS_C:
        if rx.search(name):
            return b
    return ""


def norm_text(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"\s+", " ", s).strip()


def real_url(u):
    """google.com/url-Redirect aufloesen."""
    u = htmllib.unescape(u)
    if "google.com/url" in u:
        q = parse_qs(urlparse(u).query).get("q", [u])[0]
        u = unquote(q)
    return u.strip()


def is_shop_link(u):
    try:
        host = urlparse(u).netloc.lower()
    except ValueError:
        return False
    return any(host == h or host.endswith("." + h) for h in SHOP_HOSTS)


def extract_ref(u, depth=0):
    """(platform, item_id) aus beliebiger Agent-/Shop-URL. platform: wd|tb|al."""
    if depth > 3:
        return "", ""
    try:
        p = urlparse(u)
        qs = parse_qs(p.query)
    except ValueError:
        return "", ""
    host = p.netloc.lower()
    # innere Original-URL aus Agent-Wrappern
    for k in ("url", "productLink", "goodsUrl", "product_link", "link", "goods_url"):
        if k in qs and qs[k] and unquote(qs[k][0]).startswith("http"):
            ref = extract_ref(unquote(qs[k][0]), depth + 1)
            if ref[0]:
                return ref
    # direkte Marktplatz-URLs
    if "weidian.com" in host:
        for k in ("itemID", "itemId", "item_id", "id"):
            if k in qs and qs[k][0].isdigit():
                return "wd", qs[k][0]
    if "taobao.com" in host or "tmall.com" in host:
        if "id" in qs and qs["id"][0].isdigit():
            return "tb", qs["id"][0]
    if "1688.com" in host:
        if m := re.search(r"/offer/(\d+)", p.path):
            return "al", m.group(1)
    # Agent-Formate: id-Param + Plattform-Param
    pid = next((qs[k][0] for k in ("id", "itemID", "goodsId", "item_id") if k in qs and qs[k][0].isdigit()), None)
    plat = next((qs[k][0].lower() for k in ("shop_type", "platform", "channel", "source", "shoptype") if k in qs), "")
    if pid and plat:
        if plat.startswith(("weidian", "wd")):
            return "wd", pid
        if plat.startswith(("taobao", "tb")):
            return "tb", pid
        if plat.startswith(("1688", "ali", "al")):
            return "al", pid
    # pfadbasierte Agent-Formate
    if m := re.search(r"/(?:product|item)/(weidian|taobao|tmall|1688|wd|tb|ali_1688)/(\d+)", p.path, re.I):
        g = m.group(1).lower()
        return ("wd" if g in ("weidian", "wd") else "al" if "1688" in g else "tb"), m.group(2)
    # Zahlencodes sind PRO AGENT verschieden (verifiziert Jul 2026)
    host_codes = {
        "hoobuy.com": {"1": "tb", "2": "wd", "0": "al", "3": "al"},
        "usfans.com": {"3": "wd", "2": "tb", "1": "al"},
        "oopbuy.com": {"1": "tb", "0": "al"},
    }
    if m := re.search(r"/product/(\d)/(\d+)$", p.path):
        for h, codes in host_codes.items():
            if (host == h or host.endswith("." + h)) and m.group(1) in codes:
                return codes[m.group(1)], m.group(2)
    return "", ""


def dedup_key(u):
    pf, pid = extract_ref(u)
    if pf:
        return f"{pf}:{pid}"
    try:
        p = urlparse(u)
        return f"{p.netloc.lower()}{p.path}".rstrip("/")
    except ValueError:
        return u


CUR_SYM = {"¥": "CNY", "￥": "CNY", "cny": "CNY", "rmb": "CNY", "$": "USD", "usd": "USD",
           "€": "EUR", "eur": "EUR", "£": "GBP", "gbp": "GBP"}
PRICE_VAL_RE = re.compile(
    r"([$€£¥￥]|USD|EUR|CNY|RMB|GBP)\s*(\d[\d.,]*)|(\d[\d.,]*)\s*([$€£¥￥]|USD|EUR|CNY|RMB|GBP)", re.I
)


def parse_price(p):
    """(Wert, Waehrungscode) aus Preis-Rohstring; (None, '') wenn nicht lesbar."""
    m = PRICE_VAL_RE.search(p or "")
    if not m:
        return None, ""
    cur = (m.group(1) or m.group(4) or "").strip().lower()
    num = (m.group(2) or m.group(3)).rstrip(".,")
    if "," in num and "." in num:
        if num.rfind(",") > num.rfind("."):
            num = num.replace(".", "").replace(",", ".")
        else:
            num = num.replace(",", "")
    elif "," in num:
        head, _, tail = num.rpartition(",")
        num = f"{head.replace(',', '')}.{tail}" if len(tail) <= 2 else num.replace(",", "")
    try:
        v = round(float(num), 2)
    except ValueError:
        return None, ""
    return v, CUR_SYM.get(cur, "USD")


def categorize(tab_name, section, name):
    """Prioritaet: Produkttyp im Namen > Typ in Sektion/Tab > Marken-Proxys."""
    for text in (name, section, tab_name):
        if not text:
            continue
        for cat, rx in STRONG_TYPES_C:
            if rx.search(text):
                return cat
    blob = " ".join(t for t in (name, section, tab_name) if t).lower()
    for cat, pat in CATEGORIES:
        if re.search(pat, blob):
            return cat
    return "Sonstiges"


def is_pricey(t):
    """Zelle besteht im Wesentlichen aus Preisangaben (auch mehrwaehrig mit | getrennt)."""
    if not t or not PRICE_RE.search(t):
        return False
    rest = PRICE_RE.sub("", t)
    rest = re.sub(r"[|｜/\\\s.,~\-–—()]+", "", rest)
    return len(rest) <= 3


def good_name(t, cell=None):
    if not t or len(t) < 3 or len(t) > 160:
        return False
    if cell is not None and cell.get("url"):
        return False
    if BAD_NAME.match(t) or NOISE_NAME.search(t):
        return False
    if t.startswith("http") or URL_RE.search(t):
        return False
    if re.fullmatch(r"[\d.,\s]+", t) or is_pricey(t) or PRICE_RE.fullmatch(t):
        return False
    if re.fullmatch(r"\d+[.,]?\d*\s?(g|kg|ml|cm|mm)", t, re.I):
        return False
    if re.fullmatch(r"\(?\s*\d*\s*\+?\s*(colou?r\s*ways?|colorways?|colou?rs?|styles?|versions?)\s*\)?", t, re.I):
        return False
    return True


def looks_like_header(cells_with_text):
    """Abschnitts-Header: 1-3 kurze Textzellen, keine Links, kein Preis."""
    if not (1 <= len(cells_with_text) <= 3):
        return None
    for t in cells_with_text:
        if len(t) > 40 or PRICE_RE.search(t) or NOISE_NAME.search(t):
            return None
    return " / ".join(cells_with_text)


def extract_items(grid, source_name, tab_name):
    """grid: Liste von Zeilen; Zelle = dict(text, url, img)."""
    items = []
    section = ""
    for r, row in enumerate(grid):
        texted = [c["text"] for c in row if c["text"]]
        linked = [c for c in row if c.get("url")]
        if not linked:
            hdr = looks_like_header(texted)
            if hdr:
                section = hdr
            continue
        prev = grid[r - 1] if r > 0 else []
        for idx, cell in enumerate(row):
            u = cell.get("url")
            if not u:
                continue
            # Name: Zelle selbst -> links -> rechts -> Zeile drueber (gleiche Spalte +-1)
            name = ""
            if good_name(cell["text"]):
                name = cell["text"]
            if not name:
                for j in range(idx - 1, max(-1, idx - 5), -1):
                    if 0 <= j < len(row) and good_name(row[j]["text"], row[j]):
                        name = row[j]["text"]
                        break
            if not name:
                for j in range(idx + 1, min(len(row), idx + 4)):
                    if good_name(row[j]["text"], row[j]) and not PRICE_RE.search(row[j]["text"]):
                        name = row[j]["text"]
                        break
            if not name and prev:
                for j in (idx, idx - 1, idx + 1):
                    if 0 <= j < len(prev) and good_name(prev[j]["text"], prev[j]):
                        name = prev[j]["text"]
                        break
            if not name:
                continue
            name = re.sub(r"^\d+\s*[、．.)]\s*", "", name).strip() or name
            # Preis: preisdominante Nachbarzellen, sonst Zeile drueber
            price = ""
            for j in range(max(0, idx - 4), min(len(row), idx + 5)):
                t = row[j]["text"]
                if t and t != name and is_pricey(t):
                    price = norm_text(t)[:44]
                    break
            if not price and prev:
                for j in range(max(0, idx - 2), min(len(prev), idx + 3)):
                    t = prev[j]["text"]
                    if t and t != name and is_pricey(t):
                        price = norm_text(t)[:44]
                        break
            # Bild: Zelle selbst, Nachbarn, dann Zeile drueber/drunter
            img = cell.get("img") or ""
            if not img:
                for j in range(max(0, idx - 3), min(len(row), idx + 4)):
                    if row[j].get("img"):
                        img = row[j]["img"]
                        break
            if not img:
                for adj in (prev, grid[r + 1] if r + 1 < len(grid) else []):
                    for j in range(max(0, idx - 2), min(len(adj), idx + 3)):
                        if adj[j].get("img"):
                            img = adj[j]["img"]
                            break
                    if img:
                        break
            pf, pid = extract_ref(u)
            pv, pc = parse_price(price)
            it = {
                "n": name[:120],
                "b": brand_of(name),
                "c": categorize(tab_name, section, name),
                "i": img,
                "s": source_name,
                "t": norm_text(tab_name)[:40],
            }
            if pf:
                it["pf"], it["pid"] = pf, pid
            else:
                it["u"] = u
            if pv is not None:
                it["pv"], it["pc"] = pv, pc
            elif price:
                it["p"] = price
            items.append(it)
    return items


def grid_from_html(path):
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="replace"), "lxml")
    grid = []
    for tr in soup.find_all("tr"):
        row = []
        for td in tr.find_all("td"):
            a = td.find("a", href=True)
            url = None
            if a:
                cand = real_url(a["href"])
                if is_shop_link(cand):
                    url = cand
            img = td.find("img", src=True)
            row.append({
                "text": norm_text(td.get_text(" ")),
                "url": url,
                "img": clean_img(img["src"]) if img else None,
            })
        if row:
            grid.append(row)
    return grid


def grid_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    tabs = {}
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 2:
            continue
        hyper = {}
        for hl in ws._hyperlinks:
            if hl.target:
                hyper[hl.ref] = hl.target
        grid = []
        for r, wsrow in enumerate(ws.iter_rows(), start=1):
            row = []
            for c, cell in enumerate(wsrow, start=1):
                v = cell.value
                text, url, img = "", None, None
                if v is not None:
                    v = str(v)
                    if m := HYPERLINK_RE.search(v):
                        url = htmllib.unescape(m.group(1).replace('""', '"'))
                    if m := IMAGE_RE.search(v):
                        img = clean_img(m.group(1).replace('""', '"'))
                    if not v.startswith("="):
                        text = norm_text(v)
                        if not url and (m := URL_RE.search(v)):
                            url = m.group(0)
                ref = f"{cell.coordinate}"
                if ref in hyper:
                    url = hyper[ref]
                if url:
                    url = real_url(url)
                    if not is_shop_link(url):
                        url = None
                row.append({"text": text, "url": url, "img": img})
            grid.append(row)
        tabs[ws.title] = grid
    wb.close()
    return tabs


def fetch_rates():
    """EUR-Basis-Kurse fuer Preisumrechnung; Fallback auf feste Naeherungen."""
    import requests as rq

    try:
        r = rq.get("https://api.frankfurter.app/latest?from=EUR&to=CNY,USD,GBP", timeout=15)
        rates = r.json()["rates"]
        return {"CNY": rates["CNY"], "USD": rates["USD"], "GBP": rates["GBP"], "EUR": 1.0}
    except Exception:
        return {"CNY": 7.8, "USD": 1.08, "GBP": 0.85, "EUR": 1.0}


def main():
    report = json.loads((ROOT / "data" / "fetch_report.json").read_text(encoding="utf-8"))
    names = {r["id"]: r.get("name", r["id"]) for r in report}
    tabnames = {}
    for r in report:
        for t in r.get("tabs", []):
            tabnames[(r["id"], t["gid"])] = t["name"]

    all_items = []
    per_source = {}

    for f in sorted(RAW.glob("*.html")):
        sid, gid = f.stem.rsplit("__", 1)
        src = names.get(sid, sid)
        tab = tabnames.get((sid, gid), gid)
        items = extract_items(grid_from_html(f), src, tab)
        all_items.extend(items)
        per_source[src] = per_source.get(src, 0) + len(items)

    xlsx_files = list(RAW.glob("*.xlsx")) + (list(MANUAL.glob("*.xlsx")) if MANUAL.exists() else [])
    for f in xlsx_files:
        src = names.get(f.stem, f.stem)
        for tab, grid in grid_from_xlsx(f).items():
            items = extract_items(grid, src, tab)
            all_items.extend(items)
            per_source[src] = per_source.get(src, 0) + len(items)

    # Dedup: gleiches Produkt (Plattform+ID bzw. URL) QUELLEN-UEBERGREIFEND.
    # Viele Sheets sind Klone desselben Betreibers; erste Quelle gewinnt,
    # fehlende Felder werden aus Duplikaten aufgefuellt.
    seen = {}
    unique = []
    for it in all_items:
        key = f"{it['pf']}:{it['pid']}" if "pf" in it else dedup_key(it["u"])
        if key in seen:
            prev = seen[key]
            if not prev["i"] and it["i"]:
                prev["i"] = it["i"]
            if "pv" not in prev and "pv" in it:
                prev["pv"], prev["pc"] = it["pv"], it["pc"]
                prev.pop("p", None)
            if prev["c"] == "Sonstiges" and it["c"] != "Sonstiges":
                prev["c"] = it["c"]
            if len(it["n"]) > len(prev["n"]) + 8 and not prev["b"]:
                prev["n"], prev["b"] = it["n"], it["b"]
            continue
        seen[key] = it
        unique.append(it)

    # Zweite Dedup-Stufe: exakt gleicher Name = eine Karte, QUELLEN-UEBERGREIFEND
    # (Duplikate aus verschiedenen Sheets verschmelzen; fehlende Felder auffuellen)
    seen2 = {}
    deduped = []
    for it in unique:
        key2 = re.sub(r"\s+", " ", it["n"].lower()).strip()
        if key2 in seen2:
            prev = seen2[key2]
            if not prev["i"] and it["i"]:
                prev["i"] = it["i"]
            if "pv" not in prev and "pv" in it:
                prev["pv"], prev["pc"] = it["pv"], it["pc"]
                prev.pop("p", None)
            if prev["c"] == "Sonstiges" and it["c"] != "Sonstiges":
                prev["c"] = it["c"]
            continue
        seen2[key2] = it
        deduped.append(it)
    print(f"name-dedup: {len(unique)} -> {len(deduped)}")
    unique = deduped

    # Standard-Reihenfolge: Marke -> Name; markenlose ans Ende
    unique.sort(key=lambda it: (it["b"] == "", it["b"].lower(), it["n"].lower()))

    meta = {"built": __import__("datetime").date.today().isoformat(), "rates": fetch_rates()}
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(
        json.dumps({"meta": meta, "items": unique}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"items total={len(all_items)} unique={len(unique)}  -> {OUT} ({OUT.stat().st_size // 1024} KB)")
    for src, n in sorted(per_source.items(), key=lambda x: -x[1]):
        print(f"  {src}: {n}")
    from collections import Counter
    print("Kategorien:", dict(Counter(i['c'] for i in unique).most_common()))


if __name__ == "__main__":
    main()
