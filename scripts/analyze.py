"""Struktur-Dump aller Rohdaten nach data/analysis.txt (Spalten, URL-Dichte, Beispiele)."""
import csv
import io
import re
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "analysis.txt"

URL_RE = re.compile(r"https?://\S+")


def sample_rows(rows, n=8):
    lines = []
    urlrows = 0
    for r in rows:
        if any(URL_RE.search(str(c)) for c in r if c):
            urlrows += 1
    lines.append(f"  rows={len(rows)} rows_with_url={urlrows}")
    shown = 0
    for r in rows:
        if not any(str(c).strip() for c in r if c is not None):
            continue
        cells = [str(c).strip()[:34] if c is not None and str(c).strip() else "." for c in r[:11]]
        lines.append("  | " + " ; ".join(cells))
        shown += 1
        if shown >= n:
            break
    # zusaetzlich: erste 4 Zeilen MIT URL
    shown = 0
    for r in rows:
        if any(URL_RE.search(str(c)) for c in r if c):
            cells = [str(c).strip()[:40] if c is not None and str(c).strip() else "." for c in r[:11]]
            lines.append("  U " + " ; ".join(cells))
            shown += 1
            if shown >= 4:
                break
    return lines


def main():
    out = []
    for f in sorted(RAW.glob("*.csv")):
        rows = list(csv.reader(io.StringIO(f.read_text(encoding="utf-8", errors="replace"))))
        out.append(f"===== CSV {f.name}")
        out.extend(sample_rows(rows))

    for f in sorted(RAW.glob("*.xlsx")):
        out.append(f"===== XLSX {f.name}")
        wb = openpyxl.load_workbook(f, data_only=False)
        for ws in wb.worksheets:
            out.append(f" --- tab '{ws.title}' dims={ws.dimensions} hyperlinks={len(ws._hyperlinks)}")
            rows = [[c.value for c in row] for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400))]
            out.extend(sample_rows(rows, n=5))
        wb.close()
    OUT.write_text("\n".join(out), encoding="utf-8")
    print(f"written {OUT} ({len(out)} lines)")


if __name__ == "__main__":
    main()
